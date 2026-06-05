"""Importa el Excel de RRHH (export tipo SAP) de Altas de practicantes y lo mapea
a los campos del T-Registro. Cruza por DNI con los datos que registra el trabajador.

Devuelve una lista de dicts listos para crear/actualizar registros (categoría practicante).
"""
import re
from datetime import datetime

import openpyxl

# AFP / sistema de pensiones -> código tabla 11
AFP_MAP = {
    "SNP": "02", "ONP": "02",            # Sistema Nacional de Pensiones
    "INTEGRA": "21", "HORIZONTE": "22", "PROFUTUR": "23",
    "PROFUTURO": "23", "PRIMA": "24", "HABITAT": "25",
}


def _txt(v):
    return ("" if v is None else str(v)).strip()


def _cod(v):
    """Toma el código inicial de cadenas tipo '002 - BCP' o '01 - DNI'."""
    m = re.match(r"\s*([0-9A-Za-z]+)\s*-", _txt(v))
    return m.group(1) if m else _txt(v)


def _fecha(v):
    """Normaliza fechas '20.03.2026', '20/03/2026' o datetime -> 'aaaa-mm-dd'."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = _txt(v)
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def _modalidad(clase_contrato):
    t = _txt(clase_contrato).upper()
    if "PRE PROFESIONAL" in t:
        return "2"   # Aprendizaje con predominio en CFP - prácticas pre profesionales
    if "PROFESIONAL" in t:
        return "3"   # Prácticas profesionales
    return ""


def _afp(afp_txt):
    return AFP_MAP.get(_txt(afp_txt).upper(), "")


def _tipo_doc(clase_id):
    c = _cod(clase_id)            # "01" -> "1"
    return c.lstrip("0") or c


def _sexo(tratamiento):
    t = _txt(tratamiento).upper()
    if t.startswith(("SRTA", "SRA")):
        return "2"   # femenino
    if t.startswith("SR"):
        return "1"   # masculino
    return ""


def parse(path_or_file) -> list[dict]:
    wb = openpyxl.load_workbook(path_or_file, data_only=True)
    ws = wb.active
    hdr = {}
    for c in range(1, ws.max_column + 1):
        name = _txt(ws.cell(row=1, column=c).value)
        if name:
            hdr[name] = c

    def cell(r, name):
        c = hdr.get(name)
        return ws.cell(row=r, column=c).value if c else None

    filas = []
    for r in range(2, ws.max_row + 1):
        dni = _txt(cell(r, "DNI"))
        if not dni:
            continue
        afp_cod = _afp(cell(r, "AFP"))
        clase = _txt(cell(r, "CLASE CONTRATO"))
        motivo = _txt(cell(r, "MOTIVO MEDIDA")).upper()
        # La categoría se deduce de la fila: modalidades formativas -> practicante.
        es_practicante = "FORMATIV" in motivo or "PROFESIONAL" in clase.upper()

        d = {
            "categoria": "practicante" if es_practicante else "trabajador",
            # Identidad (clave de cruce con el autoregistro del trabajador)
            "tipo_documento": _tipo_doc(cell(r, "CLASE ID")) or "1",
            "numero_documento": dni,
            # Personales (referenciales; para DNI los confirma RENIEC)
            "ap_paterno": _txt(cell(r, "PATERNO")),
            "ap_materno": _txt(cell(r, "MATERNO")),
            "nombres": _txt(cell(r, "NOMBRES")),
            "fecha_nacimiento": _fecha(cell(r, "FECHA NAC.")),
            "sexo": _sexo(cell(r, "TRATAMIENTO")),
            "email": _txt(cell(r, "EMAIL")),
            "nacionalidad": "604" if _txt(cell(r, "NACIONALIDAD")).upper() in ("PE", "PERU", "PERÚ") else "",
            # Pensión (ambas categorías)
            "regimen_pensionario": afp_cod,
            "cuspp": _txt(cell(r, "NRO. AFP")) if afp_cod in ("21", "22", "23", "24", "25") else "",
            # Banco / pago
            "entidad_bancaria": _cod(cell(r, "CLAVE BANCO")),
            "numero_cuenta": re.sub(r"\D", "", _txt(cell(r, "CUENTA BANCO"))),
            "tipo_pago": "2" if "DEPOSITO" in _txt(cell(r, "VIA PAGO")).upper() else "1",
            # Fecha de alta (inicio de vínculo / período formativo)
            "fecha_inicio_vinculo": _fecha(cell(r, "FECHA ALTA")),
            # Monto: en trabajador va al .tra (campo 16); en practicante al PLAME
            "sueldo": _txt(cell(r, "REMUNERACION")),
        }
        if es_practicante:
            d.update({
                "modalidad_formativa": _modalidad(clase),
                "seguro_medico": "1",            # EsSalud por defecto
                "discapacidad": "0",
                "tipo_centro_formacion": "2",    # Universidad (editable)
            })
        filas.append(d)
    return filas
