"""
Extrae las tablas paramétricas oficiales de SUNAT (anexo2_oficial.xlsx)
a archivos JSON que consume la app web del T-Registro.

Genera en app/data/:
  - tablas.json        : tablas simples {id: [{codigo, descripcion}, ...]}
  - ubigeo.json        : departamentos -> provincias -> distritos (cascada)
  - instituciones.json : {instituciones:[...], carreras_por_institucion:{...}}
  - ocupaciones.json   : {publico:[...], privado:[...]}

Uso:  python app/scripts/extract_tables.py
"""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "anexo2_oficial.xlsx"
OUT = ROOT / "app" / "data"
OUT.mkdir(parents=True, exist_ok=True)

# Un "código" válido: alfanumérico sin espacios (descarta títulos y notas al pie)
COD_RE = re.compile(r"^[A-Za-z0-9]+$")


def cell(v):
    return "" if v is None else str(v).strip()


def load():
    return openpyxl.load_workbook(XLSX, read_only=True, data_only=True)


# (id, hoja, col_codigo, col_descripcion)
SIMPLES = [
    ("tipo_documento", "T3- Tipo Documento", 0, 1),
    ("nacionalidad", "T4 Nacionalidad", 0, 1),
    ("via", "T5 Via", 0, 1),
    ("zona", "T6 Zona", 0, 1),
    ("tipo_trabajador", "T8 Tipo Trab-Pens-PS", 0, 1),
    ("situacion_educativa", "T9 Situación Educativa", 0, 1),
    ("modalidad_formativa", "T18 Tipo Modalidad Formativa", 0, 1),
    ("regimen_pensionario", "T11 Reg. Pensionario", 0, 1),
    ("tipo_contrato", "T12 Contratos", 0, 1),
    ("periodicidad", "T13 Periodicidad", 0, 1),
    ("eps", "T14 EPSSERV PROPIOS", 0, 2),
    ("situacion", "T15 Situación ", 0, 1),
    ("tipo_pago", "T16 Tipo de Pago", 0, 1),
    ("motivo_baja", "T17 Motivo fin del periodo", 0, 1),
    ("categoria_ocupacional", "T24 Categoria Ocupacional", 0, 1),
    ("convenio_doble_trib", "T25 Convenios", 0, 1),
    ("pais", "T26 País Emisor Dcto", 0, 1),
    ("regimen_salud", "T32 Rég Aseg Salud", 0, 1),
    ("regimen_laboral", "T33 Régimen Laboral", 0, 1),
    ("situacion_especial", "T35 Situacion especial", 0, 1),
    ("entidad_bancaria", "T36 Entidad Bancaria", 0, 1),
]


def sheet_by_hint(wb, hint):
    """Devuelve la hoja cuyo nombre coincide aunque cambien espacios/acentos."""
    if hint in wb.sheetnames:
        return wb[hint]
    norm = lambda s: re.sub(r"\s+", " ", s).strip().lower()
    for n in wb.sheetnames:
        if norm(n) == norm(hint) or norm(n).startswith(norm(hint)[:6]):
            return wb[n]
    raise KeyError(hint)


def extract_simple(wb, hoja, cc, dc):
    ws = sheet_by_hint(wb, hoja)
    out, seen = [], set()
    for row in ws.iter_rows(values_only=True):
        if max(cc, dc) >= len(row):
            continue
        cod, desc = cell(row[cc]), cell(row[dc])
        if not cod or not desc or not COD_RE.match(cod):
            continue
        if cod.upper() in ("N°", "NO", "CODIGO", "CÓDIGO", "COD"):
            continue
        if cod in seen:
            continue
        seen.add(cod)
        out.append({"codigo": cod, "descripcion": desc})
    return out


def extract_ubigeo(wb):
    ws = sheet_by_hint(wb, "T28 UBIGEO")
    # Cada columna es una lista vertical INDEPENDIENTE. La jerarquía se deduce
    # del propio código: provincia 0101 -> dpto 01; distrito 010101 -> prov 0101.
    deps, provs, dists = {}, {}, {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 6:
            continue
        cd, dd = cell(row[0]), cell(row[1])
        cp, dp = cell(row[2]), cell(row[3])
        cdi, ddi = cell(row[4]), cell(row[5])
        if len(cd) == 2 and cd.isdigit() and dd:
            deps.setdefault(cd, dd)
        if len(cp) == 4 and cp.isdigit() and dp:
            provs.setdefault(cp, {"descripcion": dp, "departamento": cp[:2]})
        if len(cdi) == 6 and cdi.isdigit() and ddi:
            dists.setdefault(cdi, {"descripcion": ddi, "provincia": cdi[:4]})
    return {
        "departamentos": [{"codigo": k, "descripcion": v} for k, v in sorted(deps.items())],
        "provincias": [{"codigo": k, **v} for k, v in sorted(provs.items())],
        "distritos": [{"codigo": k, **v} for k, v in sorted(dists.items())],
    }


def extract_instituciones(wb):
    ws = sheet_by_hint(wb, "T34 Inst. Educ. y sus carreras")
    inst, carr = {}, {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        ci, di, cc, dc = cell(row[4]), cell(row[5]), cell(row[6]), cell(row[7])
        if not (ci and ci.isdigit() and di):
            continue
        inst.setdefault(ci, di)
        if cc and dc:
            carr.setdefault(ci, [])
            if not any(x["codigo"] == cc for x in carr[ci]):
                carr[ci].append({"codigo": cc, "descripcion": dc})
    return {
        "instituciones": [{"codigo": k, "descripcion": v} for k, v in sorted(inst.items())],
        "carreras_por_institucion": carr,
    }


def extract_ocupaciones(wb):
    def grab(hoja, code_len=6):
        ws = sheet_by_hint(wb, hoja)
        out, seen = [], set()
        for row in ws.iter_rows(values_only=True):
            if len(row) < 2:
                continue
            cod, desc = cell(row[0]), cell(row[1])
            if not (cod.isdigit() and len(cod) == code_len and desc):
                continue
            if cod in seen:
                continue
            seen.add(cod)
            out.append({"codigo": cod, "descripcion": desc})
        return out

    return {
        "publico": grab("T10 Ocup SPub Pers Form"),
        "privado": grab("T30 Ocupación S.Privado"),
    }


def main():
    wb = load()
    tablas = {}
    for tid, hoja, cc, dc in SIMPLES:
        try:
            tablas[tid] = extract_simple(wb, hoja, cc, dc)
            print(f"  {tid:<24} {len(tablas[tid]):>5} filas")
        except KeyError:
            print(f"  !! no se encontró la hoja: {hoja}")

    (OUT / "tablas.json").write_text(
        json.dumps(tablas, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    ubi = extract_ubigeo(wb)
    (OUT / "ubigeo.json").write_text(json.dumps(ubi, ensure_ascii=False), encoding="utf-8")
    print(f"  ubigeo: {len(ubi['departamentos'])} dep / {len(ubi['provincias'])} prov / {len(ubi['distritos'])} dist")

    ins = extract_instituciones(wb)
    (OUT / "instituciones.json").write_text(json.dumps(ins, ensure_ascii=False), encoding="utf-8")
    print(f"  instituciones: {len(ins['instituciones'])} (carreras en {len(ins['carreras_por_institucion'])})")

    ocu = extract_ocupaciones(wb)
    (OUT / "ocupaciones.json").write_text(json.dumps(ocu, ensure_ascii=False), encoding="utf-8")
    print(f"  ocupaciones: {len(ocu['publico'])} púb / {len(ocu['privado'])} priv")

    print("OK -> archivos en", OUT)


if __name__ == "__main__":
    main()
